import tkinter
import tkinter.messagebox
import tkinter.filedialog
import customtkinter
import openpyxl
import os
import pandas as pd
import warnings
from array import array

# Modes: "System" (standard), "Dark", "Light"
customtkinter.set_appearance_mode("Dark")
# Themes: "blue" (standard), "green", "dark-blue"
customtkinter.set_default_color_theme("green")


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        #variables

        self.box_dict = {}
        self.list_excel_sheets_values = []
        self.list_excel_col_values = []
        self.list_of_cols_selected = []
        self.excelFile = []

        # configure window
        self.title("Filtro de Excel - Luisa")
        self.geometry(f"{1200}x{768}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=0)
        self.grid_columnconfigure((2, 3), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        # Sidebar

        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)

        # Sidebar content

        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="ExceLuisa", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Dark", "Light"],command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 10))
        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        # Buttons Frame

        self.buttons_frame = customtkinter.CTkFrame(self, width=300, corner_radius=10)
        self.buttons_frame.grid(row=0, column=3, rowspan=3, padx=(20, 20), pady=(100, 100), sticky="nsew")
        

        # Buttons Frame content

        self.openFileButton = customtkinter.CTkButton(self.buttons_frame, command=self.button_openFile, text="Abrir Ficheiro")
        self.openFileButton.grid(row=0, column=3, padx=(20, 20), pady=(20, 20))

        
        self.confirmSelectedColButton = customtkinter.CTkButton(self.buttons_frame, command=self.confirmSelectedCols, text="Confirmar Colunas")
        self.confirmSelectedColButton.grid(row=1, column=3, padx=(20, 20), pady=(20, 20))

        self.saveFileButton = customtkinter.CTkButton(self.buttons_frame, command=self.button_saveFile, text="Guardar Ficheiro")
        self.saveFileButton.grid(row=2, column=3, padx=(20, 20), pady=(20, 20))

        # Main Frame

        self.tab_view = customtkinter.CTkTabview(self, width=750, corner_radius=10)
        self.tab_view.grid(row=0, column=1, rowspan=3, columnspan=2, padx=(20, 20), pady=(100, 100), sticky="nsew")

        self.tab_cols = self.tab_view.add("Select Columns")  # add tab at the end
        self.tab_preview = self.tab_view.add("Preview")
        self.tab_view.set("Select Columns")
        
        #Cols Select Frame (createWidgets)


        # Preview Frame (createTextPreview)



        # Main Frame content

 

        

    ## PROBLEMA DE UI
    def createWidgets(self):
        x = 0
        y = -1
        for i in range(len(self.list_excel_col_values)):
            y = y + 1
            if(y == 5):
                x = x + 1
                y = 0
            box = customtkinter.CTkCheckBox(master=self.tab_cols, text=self.list_excel_col_values[i])
            box.grid(row=x, column=y, pady=20, padx=20)
            #box._variable
            self.box_dict[i] = box
            print(f"Option {i} is:", box._text)
       
    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)
    
    def button_openFile(self):
        self.filepathIO = tkinter.filedialog.askopenfile(mode="r", title="Abrir ficheiro Excel")
        self.filepath = self.filepathIO.name
        self.readWorkbook()
        self.createWidgets()
        
    def createTextPreview(self):
        self.textbox = customtkinter.CTkTextbox(self.tab_preview, width=600, height=300)
        self.textbox.grid(row=0, column=0)
        self.textbox.insert("0.0", text=self.excelFile)  # insert at line 0 character 0
        text = self.textbox.get("0.0", "end")  # get text from line 0 character 0 till the end
        self.textbox.configure(state="disabled")

    def confirmSelectedCols(self):
        for i in range(len(self.list_excel_col_values)):
            if(self.box_dict[i].get() != 0):
                self.list_of_cols_selected.append(self.box_dict[i].cget("text"))
        print(self.list_of_cols_selected)  
        self.updateWorkbook()
         

    def updateWorkbook(self):
        wb = pd.read_excel(self.filepath,sheet_name='data', usecols=self.list_of_cols_selected)
        self.excelFile = pd.DataFrame(wb)
        self.createTextPreview()

    def optionmenu_callback(choice):
            print("optionmenu dropdown clicked:", choice)
   
    def readWorkbook(self, ):
        wb = pd.read_excel(self.filepath,sheet_name='data')
        self.excelFile = pd.DataFrame(wb)
        for col in wb.columns:
            self.list_excel_col_values.append(col)
        
        

    def button_saveFile(self):
        self.filepathToSaveIO = tkinter.filedialog.asksaveasfile(mode="w", title="Guardar ficheiro Excel", filetypes=[("excel file", ".xlsx")], defaultextension=".xlsx")
        self.filepathToSave = self.filepathToSaveIO.name
        self.filepathToSaveIO.close()
        self.writer = pd.ExcelWriter(self.filepathToSave)
        self.excelFile.to_excel(self.writer)
        self.writer.save()
        print("Ficheiro guardado com sucesso")

if __name__ == "__main__":
    app = App()

    app.mainloop()
